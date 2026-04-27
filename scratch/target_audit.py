import openpyxl
import pandas as pd
import numpy as np

file_path = r'C:\Users\ASUS\Downloads\04 Toko-Requisition April 2026.xlsx'

def target_audit():
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    # 1. Code Sheet
    ws_code = wb['Code']
    data_code = []
    for row in ws_code.iter_rows(values_only=True):
        data_code.append(row)
    df_code = pd.DataFrame(data_code).iloc[5:]
    df_code[0] = df_code[0].astype(str).str.strip()
    
    # 2. Toko Sheet
    ws_toko = wb['Toko']
    data_toko = []
    for row in ws_toko.iter_rows(min_row=5, values_only=True):
        if len(row) > 7:
            data_toko.append({'Code': str(row[4]).strip(), 'QTY': row[7]})
    df_toko = pd.DataFrame(data_toko)
    df_toko['QTY'] = pd.to_numeric(df_toko['QTY'], errors='coerce').fillna(0)
    
    # 3. Req Sheet
    ws_req = wb['Req ']
    data_req = []
    for row in ws_req.iter_rows(min_row=4, values_only=True):
        if len(row) > 4:
            data_req.append({'Code': str(row[1]).strip(), 'QTY': row[4]})
    df_req = pd.DataFrame(data_req)
    df_req['QTY'] = pd.to_numeric(df_req['QTY'], errors='coerce').fillna(0)

    # Targeted Prefixes
    for prefix in ['02', '04']:
        print(f"\n--- Analysis for Prefix {prefix} ---")
        codes_in_summary = df_code[df_code[0].str.startswith(prefix)][0].unique()
        
        for code in codes_in_summary:
            # Summary values
            row = df_code[df_code[0] == code].iloc[0]
            val_toko_summary = pd.to_numeric(row[9], errors='coerce') or 0
            val_req_summary = pd.to_numeric(row[10], errors='coerce') or 0
            
            # Source values
            val_toko_source = df_toko[df_toko['Code'] == code]['QTY'].sum()
            val_req_source = df_req[df_req['Code'] == code]['QTY'].sum()
            
            if abs(val_toko_summary - val_toko_source) > 0.01 or abs(val_req_summary - val_req_source) > 0.01:
                print(f"Discrepancy in Code {code}:")
                print(f"  Summary: Toko={val_toko_summary}, Req={val_req_summary}")
                print(f"  Source:  Toko={val_toko_source}, Req={val_req_source}")
                print(f"  Diff:    Toko={val_toko_summary - val_toko_source}, Req={val_req_summary - val_req_source}")

target_audit()
