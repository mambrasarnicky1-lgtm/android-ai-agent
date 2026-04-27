import openpyxl
import pandas as pd

file_path = r'C:\Users\ASUS\Downloads\04 Toko-Requisition April 2026.xlsx'

def check_item_04_1004():
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws_code = wb['Code']
    data_code = []
    for row in ws_code.iter_rows(values_only=True):
        if row[0] == '04-1004':
            print(f"Code Sheet: {row}")
    
    ws_req = wb['Req ']
    qty_sum = 0
    for row in ws_req.iter_rows(min_row=4, values_only=True):
        if row[1] == '04-1004':
            print(f"Req Sheet Entry: {row}")
            qty_sum += row[4]
    print(f"Req Sheet Total QTY for 04-1004: {qty_sum}")

check_item_04_1004()
