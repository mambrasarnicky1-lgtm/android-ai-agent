import openpyxl
import pandas as pd
import numpy as np

file_path = r'C:\Users\ASUS\Downloads\04 Toko-Requisition April 2026.xlsx'

def get_sheet_data(sheet):
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return pd.DataFrame(data)

def audit():
    print("Loading workbook (read-only mode)...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    # --- 1. CODE SHEET ---
    print("Processing 'Code' sheet...")
    ws_code = wb['Code']
    df_code_raw = get_sheet_data(ws_code)
    
    cols = list(df_code_raw.iloc[3])
    row2 = list(df_code_raw.iloc[2])
    for i in range(len(cols)):
        if pd.isna(cols[i]) and not pd.isna(row2[i]):
            cols[i] = row2[i]
        cols[i] = str(cols[i]).strip()
    
    df_code = df_code_raw.iloc[5:].copy()
    df_code.columns = cols
    df_code = df_code[['Code', 'DESCRIPTION', 'Toko', 'Requisition', 'TOTAL', 'qty', 'rest']].copy()
    df_code['Code'] = df_code['Code'].astype(str).str.strip()
    for col in ['Toko', 'Requisition', 'TOTAL', 'qty', 'rest']:
        df_code[col] = pd.to_numeric(df_code[col], errors='coerce').fillna(0)
    
    # --- 2. TOKO SHEET ---
    print("Processing 'Toko' sheet...")
    ws_toko = wb['Toko']
    toko_data = []
    # Index 4 is Code, Index 7 is QTY
    for row in ws_toko.iter_rows(min_row=5, values_only=True):
        if len(row) > 7:
            toko_data.append({'Code': str(row[4]).strip(), 'QTY': row[7]})
    df_toko = pd.DataFrame(toko_data)
    df_toko['QTY'] = pd.to_numeric(df_toko['QTY'], errors='coerce').fillna(0)
    toko_summary = df_toko.groupby('Code')['QTY'].sum().reset_index()

    # --- 3. REQ SHEET ---
    print("Processing 'Req ' sheet...")
    ws_req = wb['Req ']
    req_data = []
    # Index 1 is Code, Index 4 is QTY
    for row in ws_req.iter_rows(min_row=4, values_only=True):
        if len(row) > 4:
            req_data.append({'Code': str(row[1]).strip(), 'QTY': row[4]})
    df_req = pd.DataFrame(req_data)
    df_req['QTY'] = pd.to_numeric(df_req['QTY'], errors='coerce').fillna(0)
    req_summary = df_req.groupby('Code')['QTY'].sum().reset_index()

    # --- 4. MERGE ---
    print("Analyzing discrepancies...")
    df_audit = df_code.merge(toko_summary, on='Code', how='left').rename(columns={'QTY': 'Toko_Source'})
    df_audit = df_audit.merge(req_summary, on='Code', how='left').rename(columns={'QTY': 'Req_Source'})
    
    df_audit['Toko_Source'] = df_audit['Toko_Source'].fillna(0)
    df_audit['Req_Source'] = df_audit['Req_Source'].fillna(0)
    
    df_audit['Diff_Toko'] = df_audit['Toko'] - df_audit['Toko_Source']
    df_audit['Diff_Req'] = df_audit['Requisition'] - df_audit['Req_Source']
    df_audit['Calc_Total'] = df_audit['Toko'] + df_audit['Requisition']
    df_audit['Diff_Internal_Total'] = df_audit['TOTAL'] - df_audit['Calc_Total']

    # --- 5. REPORT ---
    print("\n--- AUDIT REPORT ---\n")
    
    # 1. External discrepancies
    ext_diff = df_audit[(np.abs(df_audit['Diff_Toko']) > 0.01) | (np.abs(df_audit['Diff_Req']) > 0.01)]
    ext_diff = ext_diff[ext_diff['Code'] != 'None']
    if not ext_diff.empty:
        print(f"Discrepancies (Code vs Source Sheets): {len(ext_diff)} found.")
        print(ext_diff[['Code', 'DESCRIPTION', 'Toko', 'Toko_Source', 'Diff_Toko', 'Requisition', 'Req_Source', 'Diff_Req']].head(50).to_string())

    # 2. Internal discrepancies
    int_diff = df_audit[np.abs(df_audit['Diff_Internal_Total']) > 0.01]
    int_diff = int_diff[int_diff['Code'] != 'None']
    if not int_diff.empty:
        print(f"\nInternal Calculation Errors (Toko + Req != TOTAL): {len(int_diff)} found.")
        print(int_diff[['Code', 'DESCRIPTION', 'Toko', 'Requisition', 'TOTAL', 'Calc_Total', 'Diff_Internal_Total']].head(50).to_string())

    # 3. Focus 02 and 04
    print("\nFocus: Prefixes 02 and 04")
    for prefix in ['02', '04']:
        sub = df_audit[df_audit['Code'].str.startswith(prefix)]
        sub_err = sub[(np.abs(sub['Diff_Toko']) > 0.01) | (np.abs(sub['Diff_Req']) > 0.01) | (np.abs(sub['Diff_Internal_Total']) > 0.01)]
        print(f"\n   Prefix {prefix} (Errors: {len(sub_err)}):")
        if not sub_err.empty:
            print(sub_err[['Code', 'DESCRIPTION', 'Toko', 'Toko_Source', 'Requisition', 'Req_Source', 'Diff_Internal_Total']].to_string())

audit()
