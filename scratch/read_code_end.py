import openpyxl
import pandas as pd

file_path = r'C:\Users\ASUS\Downloads\04 Toko-Requisition April 2026.xlsx'

def read_code_end():
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['Code']
    data = []
    # Read the last 100 rows
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=max_row-100, values_only=True):
        data.append(row)
    
    df = pd.DataFrame(data)
    print("--- Code Sheet (End of sheet) ---")
    print(df.to_string())

read_code_end()
